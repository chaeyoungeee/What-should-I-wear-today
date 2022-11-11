from openpyxl import load_workbook

c_name=1
c_id=2
c_hot_level=3
c_avg_temperature=4
c_clothes_level=5

wb = load_workbook("./discordbot/userDB.xlsx", data_only=True)
ws = wb.active

def check_row():
    for row in range(2, ws.max_row+2):
        if ws.cell(row, c_name).value is None:
            return row


def check_exist(name, id):
    for row in range(2, ws.max_row+1):
        if ws.cell(row, c_name).value == name and ws.cell(row, c_id).value == hex(id):
            return False
    return True

def sign_up(name, id):
    _row = check_row()
    ws.cell(row=_row, column=c_name, value=name)
    ws.cell(row=_row, column=c_id, value=hex(id))
    ws.cell(row=_row, column=c_hot_level, value=0)
    ws.cell(row=_row, column=c_avg_temperature, value=None)
    ws.cell(row=_row, column=c_clothes_level, value=None)
    wb.save("./discordbot/userDB.xlsx")

def hot_level_update(name, id, hot_level):
    for row in range(2, ws.max_row+1):
        if ws.cell(row, c_name).value == name and ws.cell(row, c_id).value == hex(id):
            ws.cell(row=row, column=c_hot_level, value=hot_level)
            wb.save("./discordbot/userDB.xlsx")
            return

def info_update(name, id, avg_temperature, clothes_level):
    for row in range(2, ws.max_row+1):
        if ws.cell(row, c_name).value == name and ws.cell(row, c_id).value == hex(id):
            ws.cell(row=row, column=c_avg_temperature, value=avg_temperature)
            ws.cell(row=row, columnc=c_clothes_level, value=clothes_level)
            wb.save("./discordbot/userDB.xlsx")
            return

def user_save_info(name, id):
    for row in range(2, ws.max_row+1):
        if ws.cell(row, c_name).value == name and ws.cell(row, c_id).value == hex(id):
            return ws.cell(row, c_hot_level).value, ws.cell(row, c_avg_temperature).value, ws.cell(row, c_clothes_level).value
    return 0, None, None